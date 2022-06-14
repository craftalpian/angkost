import numpy as np
from tabulate import tabulate
import pyshorteners
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
import csv
import os
import glob


def shorten_url(url) -> str:
    type_tiny = pyshorteners.Shortener()

    try:
        return type_tiny.tinyurl.short(url)
    except:
        return "Tidak terdapat url"


def clear():
    files = glob.glob("./data/*")
    for f in files:
        os.remove(f)


def format_price(amount) -> int:
    return int(amount.lower().replace('rp', '').replace('.', ''))


def thousand_format(amount) -> str:
    return '{:,}'.format(amount).replace(',', '.')

def badge(badge):
    if len(badge) < 1:
        return 'Basic'
    else:
        title = badge[0]['title']
        if title == "Official Store":
            return 'Official Store'
        elif title == "Power Merchant Pro Badge":
            return 'Power Merchant Pro Badge'
        else:
            return 'Basic'

def scoring_badge(badge):
    if len(badge) < 1:
        return 1
    else:
        title = badge[0]['title']
        if title == "Official Store":
            return 5
        elif title == "Power Merchant Pro Badge":
            return 4
        else:
            return 2


def scoring_product(count_review, rating_average):
    score = float(count_review)*float(rating_average)
    return score


def sell_count(data):
    if len(data) < 1:
        return 0
    else:
        if "integrity" in str(data):
            count = str(data).split("'integrity', 'title': '")[
                1].split("'")[0].split('Terjual ')[1].split("+")[0]

            if "rb" in count:
                return int(count.split("rb")[0].strip())*1000
            else:
                return int(count.strip())
        else:
            return 0


def extract_data(data, product_name):
    data_store = []

    for i, j in enumerate(data):
        if i <= 4:
            data_store.append([j['id'], j['name'], "null", scoring_badge(j['badges']), j['countReview'], j['price'], format_price(j['price']), j['ratingAverage'],
                              j['url'], scoring_product(j['countReview'], j['ratingAverage']), scoring_product(j['countReview'], j['ratingAverage'])+scoring_badge(j['badges']), badge(j['badges']), sell_count(j['labelGroups'])])

    # Storing csv data
    with open(f"./data/result-{product_name}.csv", "w+", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(data_store)

    return data_store

# def recommendation_product(data, budget):
#     result = []

#     # Get all list of result
#     files = glob.glob("./data/*")
#     for f in files:
#         data = []
#         with open(f, 'r') as csvfile:
#             plots = csv.reader(csvfile, delimiter=',')
#             for row in plots:
#                 data.append(row)
#         result.append(data)

#     print(result)

# def recommendation_recipe(data):
#     for x in data:
#         price = 0
#         for i in x:
#             price += i[6]
#         price_avg = price/len(x)


def recommendation_system(data, budget):
    res = []
    for x in data:
        # Product-n
        prices_sum, new_product, data_product, result_product, add_point_bonus, sort_from_small = 0, [], [], [], [], []
        for i in x:
            data_product.append(i)
            prices_sum += i[6]
        avg_price = prices_sum/len(x)

        if avg_price > budget:
            new_product = x
        else:
            for j in x:
                if j[6] < avg_price:
                    new_product.append(j)

        for a in new_product:
            result_product.append([a[0], a[1], a[2], a[3], a[4], a[5],
                                  a[6], a[7], a[8], a[9], a[10], a[11], a[12], float(a[12]/a[4])])

        result_product = sorted(result_product, key=lambda x: x[13], reverse=True)
        
        for b, c in enumerate(result_product):
            sort_from_small.append([c[0], c[1], c[2], c[3], c[4], c[5], c[6], c[7], c[8],
                                   c[9], c[10], c[11], c[12], c[13], result_product[len(result_product)-1-b][13]])

        point = 0
        for k in sort_from_small:
            point += k[14]
        avg_point = point/len(sort_from_small)

        for u in sort_from_small:
            bonus = 0
            if u[11] == "Official Store":
                bonus = avg_point*1.0
            elif u[11] == "Power Merchant Pro Badge":
                bonus = avg_point*0.75
            else:
                bonus = avg_point*0.5

            add_point_bonus.append([u[0], u[1], u[2], u[3], u[4], u[5], u[6], u[7], u[8], u[9], u[10], u[11], u[12], u[13], u[14]+bonus])
        last = sorted(add_point_bonus, key=lambda x: x[14], reverse=True)

        res.append(last[0])
    return res


def final_result(data, data_result, budget):
    # Storage (by price - LOW)
    stored_price_low = {}

    # Storage (by price - HIGH)
    stored_price_high = {}

    # Storage (by score - LOW)
    stored_score_low = {}

    # Storage (by score - HIGH)
    stored_score_high = {}

    rec_system = recommendation_system(data_result, budget)

    for x, y in enumerate(data_result):
        stored_price_low[data[x]] = sorted(
            y, key=lambda x: x[6], reverse=False)
        stored_price_high[data[x]] = sorted(
            y, key=lambda x: x[6], reverse=True)
        stored_score_low[data[x]] = sorted(
            y, key=lambda x: x[10], reverse=False)
        stored_score_high[data[x]] = sorted(
            y, key=lambda x: x[10], reverse=True)

    # Result
    result = {}

    # Sort by price - low
    sort_by_price_low = []
    # Sort by price - high
    sort_by_price_high = []
    # Sort by score - low
    sort_by_score_low = []
    # Sort by score - high
    sort_by_score_high = []

    for j in data:
        # cheap
        sort_by_price_low.append(stored_price_low[j][0])
        sort_by_price_high.append(stored_price_high[j][0])
        # score recommendation
        sort_by_score_low.append(stored_score_low[j][0])
        sort_by_score_high.append(stored_score_high[j][0])

    result["price_low"] = sort_by_price_low
    result["price_high"] = sort_by_price_high
    result["score_low"] = sort_by_score_low
    result["score_high"] = sort_by_score_high
    result["recommendation"] = rec_system

    wb = Workbook()

    # Worksheet 1
    ws1 = wb.active
    ws1.title = "LOW_PRICE (ANGKOST)"
    ws1.sheet_properties.tabColor = "83A9AC"

    ws1.append(["Id Produk", "Nama Produk", "Harga Produk",
               "Bintang Produk", "Link Produk", "Skor Produk"])

    for i in sort_by_price_low:
        ws1.append([i[0], i[1], i[5], i[7], i[8], i[9]])

    ws2 = wb.create_sheet("HIGH_PRICE (ANGKOST)")

    ws2.sheet_properties.tabColor = "D4D39B"

    ws2.append(["Id Produk", "Nama Produk", "Harga Produk",
               "Bintang Produk", "Link Produk", "Skor Produk"])

    for i in sort_by_price_high:
        ws2.append([i[0], i[1], i[5], i[7], i[8], i[9]])

    ws3 = wb.create_sheet("LOW_SCORE (ANGKOST)")

    ws3.sheet_properties.tabColor = "C9A992"

    ws3.append(["Id Produk", "Nama Produk", "Harga Produk",
               "Bintang Produk", "Link Produk", "Skor Produk"])

    for i in sort_by_score_low:
        ws3.append([i[0], i[1], i[5], i[7], i[8], i[9]])

    ws4 = wb.create_sheet("HIGH_SCORE (ANGKOST)")

    ws4.sheet_properties.tabColor = "242A30"

    ws4.append(["Id Produk", "Nama Produk", "Harga Produk",
               "Bintang Produk", "Link Produk", "Skor Produk"])

    for i in sort_by_score_high:
        ws4.append([i[0], i[1], i[5], i[7], i[8], i[9]])

    ws5 = wb.create_sheet("RECOMMENDATION (ANGKOST)")

    ws5.sheet_properties.tabColor = "FFC400"

    ws5.append(["Id Produk", "Nama Produk", "Harga Produk",
               "Bintang Produk", "Link Produk", "Skor Produk (Rekomendasi)"])

    for i in rec_system:
        ws5.append([i[0], i[1], i[5], i[7], i[8], i[14]])

    wb.save(filename="angkost_result.xlsx")

    for i in result:
        table_detail(i, result[i], budget)

    return result


def short(text, amount=50):
    if len(text) > amount:
        return text[:amount]+"..."
    else:
        return text


def table_detail(title, data, budget):
    print(f"\n[{title.upper()}]:")
    data_new = [[x+1, y[1], y[5], y[4], y[7], y[10]]
                for x, y in enumerate(data)]
    print(tabulate(data_new, headers=[
          "#", "nama" "harga", "penjualan", "bintang", "skor"]))
    price_total = np.sum(np.array([x[6] for x in data]))
    print(f"\nTotal: Rp{thousand_format(price_total)}\t\tBudget: Rp{thousand_format(budget)}\t\tSelisih: Rp{thousand_format(budget-price_total)}\n")
